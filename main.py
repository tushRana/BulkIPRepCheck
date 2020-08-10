# -*- coding: utf-8 -*-
"""
Created on Sun Jun 21 13:51:24 2020

@author: tusrana
"""
import os
import sys
import argparse
import pandas as pd
import ipaddress
import requests
import urllib3
from requests.auth import HTTPBasicAuth 
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill

################COLORCODING###################
G = '\033[92m'  # green
Y = '\033[93m'  # yellow
B = '\033[94m'  # blue
R = '\033[91m'  # red
W = '\033[0m'   # white

################BANNER###################
def banner():
    print("  ____        _ _    ___ ____  ____             ____ _               _    ")
    print(" | __ ) _   _| | | _|_ _|  _ \|  _ \ ___ _ __  / ___| |__   ___  ___| | __")
    print(" |  _ \| | | | | |/ /| || |_) | |_) / _ \ '_ \| |   | '_ \ / _ \/ __| |/ /")
    print(" | |_) | |_| | |   < | ||  __/|  _ <  __/ |_) | |___| | | |  __/ (__|   < ")
    print(" |____/ \__,_|_|_|\_\___|_|   |_| \_\___| .__/ \____|_| |_|\___|\___|_|\_\ ")
    print("                                        |_|                               ")
    print(B + "                                                         By - @oldsoul     " + W)

################PARSE ARGUMENTS###################
def parser_error(errmsg):
    banner()
    print("Usage: python " + sys.argv[0] + " [Options] use -h for help")
    print(R + "Error: " + errmsg + W)
    sys.exit()

def parse_args():
    # parse the arguments
    parser = argparse.ArgumentParser(prog='BulkIPRepCheck', formatter_class=argparse.RawDescriptionHelpFormatter, epilog='Example: \n python ' + sys.argv[0] + ' -f /path/to/file \n python ' + sys.argv[0] + '-f /path/to/file -s xforce,alienvault -o outfile.xlsx')
    parser.error = parser_error
    parser._optionals.title = "OPTIONS"
    parser.add_argument('-f', '--file', help="Input file containing IPs", required=False)
    #parser.add_argument('-v', '--verbose', help='Enable Verbosity and display results in realtime', action="store_true")
    parser.add_argument('-t', '--threads', help='Number of threads to use for subbrute bruteforce', type=int, default=5)
    parser.add_argument('-s', '--sources', help='Specify a comma-separated list of sources to query')
    parser.add_argument('-sL', '--list_sources', help='List all available sources', action="store_true")
    parser.add_argument('-o', '--output', help='Save the results to text file')
    parser.add_argument('--version', action='version', version='%(prog)s v1.1')
    return parser.parse_args()

################IP REGEX CHECK###################
def is_ipv4(string):
    try:
        ipaddress.IPv4Network(string)
        return True
    except ValueError:
        return False

##################READ API KEYS##################
def read_keys():
    try:
        f = open("api_keys.txt", "r")
    except:
        print ("API Keys file not found")
        exit()
    else:
        username_line = f.readline()
        username = username_line.split('=', 1)
        username = re.sub(r"[\n\t\s]*", "", username[1])
        password_line = f.readline()
        password = password_line.split('=', 1)
        password = re.sub(r"[\n\t\s]*", "", password[1])
        if not username:
            return "false"
        if not password:
            return "false"
        credentials = [username,password]
        return credentials

##################READ API KEYS##################
def read_input(file_path):
    if file_path and os.path.exists(file_path):
        print ("[*] Parsing Source IPs")
        count = 0
        IP_list = []
        if file_path.endswith('.xlsx'):
            IP_data = pd.read_excel (file_path)
            for j in range(len(IP_data.columns)):
                for i in range(len(IP_data.index)):
                    temp = str(IP_data.loc[i][j])
                    if (is_ipv4(temp)):
                        count=count+1
                        IP_list.append(temp)
        elif file_path.endswith('.txt'):
            file = open(file_path, 'r')
            IP_list = file.readlines()
            count = len(IP_list)
        print(Y + " [+] Total IPs:" + W, count)
        return IP_list
    else:
        sys.exit(R + "[!] Source File not found." + W)

##################CREATE OUTPUT FILE##################
def create_output_file(IP_list, output):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reputations"
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name='Verdana', size=9, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FFFFFF')
    highlight.fill = PatternFill(fill_type="solid", bgColor='FF000000')
    workbook.add_named_style(highlight)
    cell_obj = sheet.cell(row = 1, column = 1)
    cell_obj.value = "IP"
    sheet.merge_cells('A1:A2')
    sheet['A1'].style = highlight
    row = 3
    col = 1
    for IP in IP_list:
        cell_obj = sheet.cell(row = row, column = col)
        cell_obj.value = IP
        row = row + 1

    workbook.save(output + ".xlsx")

#################IBM XFORCE API##################
def fetch_xforce(IP_list, output):
    print(G + "[*] Querying IBM XFORCE" + W)
    try:
        f = open("api_keys.txt", "r")
    except:
        print (Y + " [+] API Keys file not found. Moving on..." + W)
        return
    else:
        username_line = f.readline()
        username = username_line.split('=', 1)
        username = re.sub(r"[\n\t\s]*", "", username[1])
        password_line = f.readline()
        password = password_line.split('=', 1)
        password = re.sub(r"[\n\t\s]*", "", password[1])
        if not username or not password:
            print(Y + " [+] IBM XFORCE API not configured. Moving on..." + W)
            return
        creds = [username,password]
    
    BASE_URL = "https://api.xforce.ibmcloud.com/ipr/history/"
    wb_obj = openpyxl.load_workbook(output + ".xlsx")
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 1, column = 2)
    cell_obj.value = "IBM XFORCE"
    sheet_obj.merge_cells('B1:D1')
    sheet_obj['B1'].style = 'highlight'
    cell_obj = sheet_obj.cell(row = 2, column = 2)
    cell_obj.value = "Status"
    cell_obj = sheet_obj.cell(row = 2, column = 3)
    cell_obj.value = "Risk Score"
    cell_obj = sheet_obj.cell(row = 2, column = 4)
    cell_obj.value = "Malware Instances"
    row = 3
    col = 2
    
    for IP in IP_list:
        url = BASE_URL + IP
        try:
            response = requests.get(url = url , auth = HTTPBasicAuth(creds[0], creds[1]))
        except:
            print(R + " [!] Cannot connect to the API. Moving on..." + W)
            return
        else:
            if response.status_code == 403:
                issue = "Invalid API keys"
                cell_obj = sheet_obj.cell(row = row, column = col)
                cell_obj.value = issue
            elif response.status_code == 401:
                issue = "Execution Error"
                cell_obj = sheet_obj.cell(row = row, column = col)
                cell_obj.value = issue
            elif response.status_code == 404:
                issue = "No record found"
                cell_obj = sheet_obj.cell(row = row, column = col)
                cell_obj.value = issue
            elif response.status_code == 200:
                response_json = response.json()
                ip_history = response_json['history']
                #df = pd.DataFrame(ip_history).to_excel("excel.xlsx")
                seen_in = []
                risk_score = 0
                count1 = 0
                for trav in ip_history:
                    risk_score = risk_score + trav['score']
                    count1 = count1 + 1
                    if 'malware_extended' in trav:
                        malware = trav['malware_extended']
                        if malware['BotNet'] not in seen_in:
                            seen_in.append(malware['BotNet'])
                risk_score_avg = risk_score/count1
                if risk_score_avg > 5:
                    status = "Malicious"
                else:
                    status = "Not-Malicious"
                cell_obj = sheet_obj.cell(row = row, column = col)
                cell_obj.value = status
                cell_obj = sheet_obj.cell(row = row, column = col+1)
                cell_obj.value = risk_score_avg
                cell_obj = sheet_obj.cell(row = row, column = col+2)
                cell_obj.value = str(seen_in)
            row = row+1
    wb_obj.save(output + ".xlsx")

#####################ABUSE.CH####################
def fetch_abusech(IP_list, output):
    print(G + "[*] Querying ABUSE.CH" + W)
    botnet_url = "https://feodotracker.abuse.ch/downloads/ipblocklist.csv"
    wb_obj = openpyxl.load_workbook(output + ".xlsx")
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 1, column = 5)
    cell_obj.value = "abuse.ch"
    sheet_obj.merge_cells('E1:F1')
    sheet_obj['E1'].style = 'highlight'
    cell_obj = sheet_obj.cell(row = 2, column = 5)
    cell_obj.value = "Status"
    cell_obj = sheet_obj.cell(row = 2, column = 6)
    cell_obj.value = "Found-In"
    row = 3
    col = 5
    
    try:
        response = requests.get(url = botnet_url)
    except:
        print(R + " [!] Cannot connect to the server. Moving on..." + W)
        return
    res_data = response.text
    res_data = res_data.split("\n")
    res_data = list(filter(None, res_data))
    res_data_sntised = []
    arr = []
    for i in range(0, len(res_data)):
        if res_data[i][0]!="#":
            res_data_sntised.append(res_data[i])
    for i in res_data_sntised:
        temp = i.split(",", 5)
        col1=[]
        col1.append(temp[1])
        col1.append(temp[4])
        arr.append(col1) 
    
    for IP in IP_list:
        flag = 0
        for ips in arr:
            if ips[0]==IP:
                flag = 1
                found_in = ips[1]
        if flag == 1:
            cell_obj = sheet_obj.cell(row = row, column = col)
            cell_obj.value = "Malicious"
            cell_obj = sheet_obj.cell(row = row, column = col+1)
            cell_obj.value = found_in
        else:
            cell_obj = sheet_obj.cell(row = row, column = col)
            cell_obj.value = "Not-Malicious"
        row = row + 1
    
    wb_obj.save(output + ".xlsx")

##############CISCO TALOS########################
def fetch_talos(IP_list, output):
    print(G + "[*] Querying CISCO Talos" + W)
    blacklist_url = "https://talosintelligence.com/documents/ip-blacklist"
    wb_obj = openpyxl.load_workbook(output + ".xlsx")
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 1, column = 7)
    cell_obj.value = "CISCO Talos"
    sheet_obj['G1'].style = 'highlight'
    cell_obj = sheet_obj.cell(row = 2, column = 7)
    cell_obj.value = "Status"
    row = 3
    col = 7
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    try:
        response = requests.get(url = blacklist_url, verify=False)
    except:
        print(R + " [!] Cannot connect to the server. Moving on..." + W)
        return
    res_data = response.text
    res_data = res_data.split("\n")
    res_data = list(filter(None, res_data))
    
    for IP in IP_list:
        flag = 0
        for ips in res_data:
            if ips==IP:
                flag = 1
        if flag == 1:
            cell_obj = sheet_obj.cell(row = row, column = col)
            cell_obj.value = "Malicious"
        else:
            cell_obj = sheet_obj.cell(row = row, column = col)
            cell_obj.value = "Not-Malicious"
        row = row + 1
    
    wb_obj.save(output + ".xlsx")
    
#################ALIENVAULT######################
def fetch_alienvault(IP_list, output):
    print(G + "[*] Querying Alienvault" + W)
    url = "https://reputation.alienvault.com/reputation.generic"
    wb_obj = openpyxl.load_workbook(output + ".xlsx")
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 1, column = 8)
    cell_obj.value = "Alienvault"
    sheet_obj['H1'].style = 'highlight'
    cell_obj = sheet_obj.cell(row = 2, column = 8)
    cell_obj.value = "Status"
    row = 3
    col = 8
    
    try:
        response = requests.get(url = url)
    except:
        print(R + " [!] Cannot connect to the server. Moving on..." + W)
        return
    res_data = response.text
    res_data = res_data.split("\n")
    res_data = list(filter(None, res_data))
    res_data_sntised = []
    arr = []
    for i in range(0, len(res_data)):
        if res_data[i][0]!="#":
            res_data_sntised.append(res_data[i])
    for i in res_data_sntised:
        temp = i.split(" # ", 2)
        col1=[]
        col1.append(temp[0])
        col1.append(temp[1])
        arr.append(col1) 
    
    for IP in IP_list:
        flag = 0
        for ips in arr:
            if ips[0]==IP:
                flag = 1
        if flag == 1:
            cell_obj = sheet_obj.cell(row = row, column = col)
            cell_obj.value = "Malicious"
        else:
            cell_obj = sheet_obj.cell(row = row, column = col)
            cell_obj.value = "Not-Malicious"
        row = row + 1
    
    wb_obj.save(output + ".xlsx")
    
#################EXECUTE SOURCES######################
def run_sources(sources, IP_list, output):
    if sources is None:
        fetch_xforce(IP_list, output)
        fetch_abusech(IP_list, output)
        fetch_talos(IP_list, output)
        fetch_alienvault(IP_list, output)
    else:
        sources = sources.split(',')
        for source in sources:
            if source.lower() == 'xforce':
                fetch_xforce(IP_list, output)
            if source.lower() == 'abusech':
                fetch_abusech(IP_list, output)
            if source.lower() == 'talos':
                fetch_talos(IP_list, output)
            if source.lower() == 'alienvault':
                fetch_alienvault(IP_list, output)

####################MAIN#########################
def main():
    #Parse Arguments
    args = parse_args()
    file = args.file
    #verbose = args.verbose
    threads = args.threads
    sources = args.sources
    list_sources = args.list_sources
    output = args.output
    if list_sources:
        print("Supported Sources: xforce,abusech,talos,alienvault")
        exit()
    if output:
        output = output.split('.')
    elif not output:
        output = "IP_Reputations.xlsx"
        
    #Display Banner
    banner()
    #Read Input File
    IP_list = read_input(file)
    #Create Output File
    create_output_file(IP_list, output[0])

    #Running tests
    run_sources(sources, IP_list, output[0])

if __name__ == "__main__":
    main()